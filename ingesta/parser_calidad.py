"""Parser de los reportes de calidad Binlab ('Evaluación de Calidad').

Replica frmInicio.CommandButton21 del VBA:
  - Lee la hoja 'Evaluación de Calidad'.
  - El no_cargue está en R5C18 (fila 5, columna 18).
  - Tres bloques de defectos: menores (filas 19→20, cols causal/%/incidencia
    3/7/6), mayores (19→30, cols 10/14/13), críticos (19→21, cols 17/21/20).
  - Cada causal genera una fila con causal + porcentaje.

Devuelve dataclasses; la persistencia (cruce con kg_consolidado para traza,
zona, lote, kg nacional) la hace el servicio de causales.
"""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class CausalFila:
    no_cargue: int
    fecha: dt.date | None
    causal: str
    porcentaje: float
    severidad: str  # 'MENOR' | 'MAYOR' | 'CRITICO'


@dataclass
class ReporteCalidad:
    archivo: str
    no_cargue: int | None
    fecha: dt.date | None
    causales: list[CausalFila]
    avisos: list[str]


# (severidad, col_causal, col_porcentaje, fila_fin)
BLOQUES = [
    ("MENOR", 3, 7, 20),
    ("MAYOR", 10, 14, 30),
    ("CRITICO", 17, 21, 21),
]


def _num(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _fecha(v) -> dt.date | None:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def parsear(ruta: str | Path) -> ReporteCalidad:
    ruta = Path(ruta)
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
    avisos: list[str] = []

    nombre = next((s for s in wb.sheetnames if "evaluaci" in s.lower() and "calidad" in s.lower()), None)
    if not nombre:
        wb.close()
        return ReporteCalidad(ruta.name, None, None, [], [f"No encontré hoja 'Evaluación de Calidad' en {ruta.name}"])
    ws = wb[nombre]

    no_cargue = None
    v = ws.cell(5, 18).value  # R5C18
    try:
        no_cargue = int(float(v)) if v not in (None, "") else None
    except (TypeError, ValueError):
        avisos.append(f"no_cargue en R5C18 no numérico: {v!r}")

    fecha = _fecha(ws.cell(5, 4).value)

    causales: list[CausalFila] = []
    for severidad, col_causal, col_pct, fila_fin in BLOQUES:
        for r in range(19, fila_fin):
            causal = ws.cell(r, col_causal).value
            pct = _num(ws.cell(r, col_pct).value)
            if causal in (None, "") or pct <= 0:
                continue
            causales.append(CausalFila(
                no_cargue=no_cargue or 0,
                fecha=fecha,
                causal=str(causal).strip(),
                porcentaje=pct,
                severidad=severidad,
            ))
    wb.close()
    return ReporteCalidad(ruta.name, no_cargue, fecha, causales, avisos)
