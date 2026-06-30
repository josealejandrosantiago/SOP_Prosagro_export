"""Parser del 'Informe de proceso semana N' que envía la maquila Binlab.

Reemplaza al script externo `procesar_informe_proceso.py` que el usuario
corre hoy a mano. Cambios:
  - No reescribe el archivo origen — devuelve estructuras Python que se
    insertan en la BD.
  - El ajuste por calibre N/A se marca con categoria='AJUSTE' (no se pierde).
  - La compensación contraria automática en fruta nacional se aplica en el
    motor SOP, no acá (acá solo parseamos).

Reglas de transformación que mantiene:
  - consec_int (122/123/124 + lote) → zona_interna (02/01/03) + lote zfill(2).
  - Trazabilidad: '2026 0MM cargue zona lote'.
  - Calibre 26 (EUR26) sale de export y se trata como 'SIMULACION' en
    fruta_nacional (categoria adicional).
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl


# ───────────────────────── Constantes del dominio ────────────────────────────
ZONA_EXT_A_INT = {"122": "02", "123": "01", "124": "03", "125": "04", "127": "06"}

# ───────────────────────── Dataclasses de salida ─────────────────────────────
@dataclass
class Ingreso:
    trazabilidad: str
    semana: int
    anio: int
    fecha_ingreso: dt.date
    no_cargue: int
    zona: str
    lote: str
    consec_int: str
    canastillas: int
    peso_neto: float
    conductor: str | None
    placa: str | None
    finaliza: str | None


@dataclass
class FrutaExportFila:
    trazabilidad: str
    semana: int
    anio: int
    dia_sem: str | None
    fecha_ingreso: dt.date
    fecha_procesamiento: dt.date
    no_cargue: int
    presentacion_caja: str | None
    calibre_desc: str | None
    calibre_num: str | None
    id_calibre: str | None
    cant_cajas: float
    total_kg_netos: float
    productor_nombre: str | None
    producto: str | None
    ica: str | None
    ggn: str | None
    predio: str | None
    categoria: str = "C1"        # 'C1' | 'C2' | 'AJUSTE'


@dataclass
class FrutaNacionalFila:
    trazabilidad: str
    semana: int
    anio: int
    dia_sem: str | None
    fecha_ingreso: dt.date
    fecha_procesamiento: dt.date
    no_cargue: int
    lote_proceso: str | None
    merma: float
    cant_kilos_descarte: float
    simulacion_kg: float = 0.0


@dataclass
class InformeMaquila:
    semana: int
    anio: int
    ingresos: list[Ingreso] = field(default_factory=list)
    export: list[FrutaExportFila] = field(default_factory=list)
    nacional: list[FrutaNacionalFila] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)


# ───────────────────────── Helpers ───────────────────────────────────────────
def _solo_fecha(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(v))
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return dt.date(y, mo, d)


def _transformar_consec(v) -> tuple[str | None, str | None, str | None]:
    """('124 101') -> (zona_interna='03', lote='101', consec_normalizado='03 101')."""
    if v is None:
        return None, None, None
    s = str(v).strip()
    m = re.match(r"^(\d+)\s+(\w+)$", s)
    if not m:
        return None, None, None
    zona_ext, lote = m.group(1), m.group(2)
    zona_int = ZONA_EXT_A_INT.get(zona_ext, zona_ext)
    if lote.isdigit():
        lote = lote.zfill(2)
    return zona_int, lote, f"{zona_int} {lote}"


def _trazabilidad(anio: int, fecha: dt.date, no_cargue: int, consec_norm: str) -> str:
    """2026 00M cargue 'zona lote'  (M = mes del ingreso)."""
    return f"{anio} 0{fecha.month:02d} {no_cargue} {consec_norm}"


def _leer_hoja(ws) -> list[dict]:
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        rows.append({h: ws.cell(r, col).value for h, col in hdr.items() if h is not None})
    return rows


def _es_calibre_26(row: dict) -> bool:
    """MID(id_calibre, 4, 2) == '26' — la regla del script original."""
    idc = str(row.get("id_calibre", "") or "")
    return idc[3:5] == "26"


def _es_calibre_na(row: dict) -> bool:
    """Calibre N/A → ajuste administrativo."""
    cal = str(row.get("calibre", "") or "").upper()
    return "N/A" in cal


def _get_anio(rows: list[dict]) -> int:
    """Devuelve el año más frecuente entre las fechas, o el actual."""
    años = []
    for r in rows:
        f = _solo_fecha(r.get("fecha"))
        if f:
            años.append(f.year)
    if not años:
        return dt.date.today().year
    return max(set(años), key=años.count)


# ───────────────────────── API pública ───────────────────────────────────────
def parsear(ruta: str | Path) -> InformeMaquila:
    """Lee el xlsx CRUDO de la maquila y devuelve el informe estructurado.

    El xlsx debe tener al menos las hojas 'Informe de Ingreso', 'export' (o
    'Export') y 'Nal'.
    """
    ruta = Path(ruta)
    wb = openpyxl.load_workbook(ruta, data_only=True)

    nombre_export = "export" if "export" in wb.sheetnames else "Export"
    ing_rows = _leer_hoja(wb["Informe de Ingreso"])
    exp_rows = _leer_hoja(wb[nombre_export])
    nal_rows = _leer_hoja(wb["Nal"])

    anio = _get_anio(ing_rows or exp_rows)
    semana_val = next(
        (int(r["semana"]) for r in ing_rows if r.get("semana")),
        next((int(r["semana"]) for r in exp_rows if r.get("semana")), 0),
    )

    informe = InformeMaquila(semana=semana_val, anio=anio)

    # ───── Ingreso ──────────────────────────────────────────────────────────
    for r in ing_rows:
        fecha = _solo_fecha(r.get("fecha"))
        if fecha is None:
            informe.avisos.append(f"Ingreso sin fecha — fila omitida: {r}")
            continue
        zona_int, lote, consec_norm = _transformar_consec(r.get("consec_int"))
        if not consec_norm:
            informe.avisos.append(f"consec_int inválido — fila omitida: {r}")
            continue
        no_cargue = int(r["no_cargue"])
        traza = _trazabilidad(anio, fecha, no_cargue, consec_norm)
        informe.ingresos.append(
            Ingreso(
                trazabilidad=traza,
                semana=int(r["semana"]),
                anio=anio,
                fecha_ingreso=fecha,
                no_cargue=no_cargue,
                zona=zona_int,
                lote=lote,
                consec_int=consec_norm,
                canastillas=int(r.get("canastillas") or 0),
                peso_neto=float(r.get("peso_neto") or 0),
                conductor=(r.get("cond_nombres") or None),
                placa=(r.get("veh_placa") or None),
                finaliza=(r.get("finaliza") or None),
            )
        )

    # ───── Export (filtramos calibre 26 → simulación; N/A → ajuste) ─────────
    sim_kg_por_traza: dict[str, float] = {}
    for r in exp_rows:
        fecha = _solo_fecha(r.get("fecha"))
        if fecha is None:
            informe.avisos.append(f"Export sin fecha — fila omitida: {r}")
            continue
        zona_int, lote, consec_norm = _transformar_consec(r.get("consec_int"))
        if not consec_norm:
            informe.avisos.append(f"Export con consec_int inválido — omitida: {r}")
            continue
        no_cargue = int(r["no_cargue"])
        traza = _trazabilidad(anio, fecha, no_cargue, consec_norm)
        cant_cajas = float(r.get("cant_cajas") or 0)
        total_kg = float(r.get("total_kilos_netos") or 0)
        calibre_num = None
        idcal = str(r.get("id_calibre") or "")
        if len(idcal) >= 5:
            calibre_num = idcal[3:5]

        if _es_calibre_26(r):
            sim_kg_por_traza[traza] = sim_kg_por_traza.get(traza, 0.0) + total_kg
            continue  # NO entra a fruta_export

        categoria = "AJUSTE" if _es_calibre_na(r) else "C1"
        if calibre_num == "N/":
            calibre_num = "N/A"

        informe.export.append(
            FrutaExportFila(
                trazabilidad=traza,
                semana=int(r["semana"]),
                anio=anio,
                dia_sem=(r.get("dia_sem") or None),
                fecha_ingreso=fecha,
                fecha_procesamiento=fecha,  # la maquila a veces no separa, lo cuadra el motor
                no_cargue=no_cargue,
                presentacion_caja=(r.get("caja") or None),
                calibre_desc=(r.get("calibre") or None),
                calibre_num=calibre_num,
                id_calibre=(r.get("id_calibre") or None),
                cant_cajas=cant_cajas,
                total_kg_netos=total_kg,
                productor_nombre=(r.get("nombres") or None),
                producto=(r.get("producto") or None),
                ica=str(r.get("ICA") or "") or None,
                ggn=str(r.get("GGN") or "") or None,
                predio=(r.get("nombre_finca") or None),
                categoria=categoria,
            )
        )

    # ───── Nacional ─────────────────────────────────────────────────────────
    for r in nal_rows:
        fecha = _solo_fecha(r.get("fecha"))
        if fecha is None:
            informe.avisos.append(f"Nacional sin fecha — fila omitida: {r}")
            continue
        zona_int, lote, consec_norm = _transformar_consec(r.get("consec_int"))
        if not consec_norm:
            informe.avisos.append(f"Nacional con consec_int inválido — omitida: {r}")
            continue
        no_cargue = int(r["no_cargue"])
        traza = _trazabilidad(anio, fecha, no_cargue, consec_norm)
        informe.nacional.append(
            FrutaNacionalFila(
                trazabilidad=traza,
                semana=int(r.get("no_semana") or r.get("semana") or 0),
                anio=anio,
                dia_sem=(r.get("dia_sem") or None),
                fecha_ingreso=fecha,
                fecha_procesamiento=fecha,
                no_cargue=no_cargue,
                lote_proceso=(r.get("lote") or None),
                merma=float(r.get("merma") or 0),
                cant_kilos_descarte=float(r.get("cant_kilos_descarte") or 0),
                simulacion_kg=sim_kg_por_traza.get(traza, 0.0),
            )
        )

    # Trazabilidades del calibre 26 que no aparecen en Nal — las añadimos sólo
    # como fila virtual con descarte=0 y simulación>0.
    trazas_nal = {f.trazabilidad for f in informe.nacional}
    for traza, sim_kg in sim_kg_por_traza.items():
        if traza not in trazas_nal:
            ing = next((i for i in informe.ingresos if i.trazabilidad == traza), None)
            if not ing:
                continue
            informe.nacional.append(
                FrutaNacionalFila(
                    trazabilidad=traza,
                    semana=ing.semana,
                    anio=ing.anio,
                    dia_sem=None,
                    fecha_ingreso=ing.fecha_ingreso,
                    fecha_procesamiento=ing.fecha_ingreso,
                    no_cargue=ing.no_cargue,
                    lote_proceso=None,
                    merma=0,
                    cant_kilos_descarte=0,
                    simulacion_kg=sim_kg,
                )
            )

    return informe


if __name__ == "__main__":  # pragma: no cover
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Uso: python -m ingesta.parser_maquila <archivo.xlsx>")
    inf = parsear(sys.argv[1])
    print(json.dumps(
        {
            "semana": inf.semana,
            "anio": inf.anio,
            "ingresos": len(inf.ingresos),
            "export": len(inf.export),
            "nacional": len(inf.nacional),
            "avisos": inf.avisos[:10],
        },
        indent=2,
        default=str,
    ))
