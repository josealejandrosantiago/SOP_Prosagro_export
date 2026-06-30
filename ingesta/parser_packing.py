"""Parser de los packing lists de Prosagro.

Soporta dos formatos:
  1) **Legacy** (`PACKING LIST EXP #326.xlsx`): tres hojas separadas
     'RESUMEN CONTENEDOR' / 'DETALLE POR PALLET' / 'insumos'.
  2) **Nuevo TNLC** (`PACKING LIST TNLC-313.xlsx`): UNA sola hoja con header
     en R8 (PALET N° / CALIBRE / PREDIO / GGN / REG. ICA / CLIENTE / Total /
     CERTIFICADO) y bloque de clientes/VAT a la derecha (cols J-N).

Devuelve estructuras Python listas para insertar en BD.
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


# ───────────────────────── Dataclasses ───────────────────────────────────────
@dataclass
class PalletDetalle:
    no_pallet: int
    calibre: str
    presentacion_caja: str | None
    predio: str | None
    ica: str | None
    ggn: str | None
    no_cargue: int
    cajas: float
    total_cajas_pallet: float | None = None  # solo en la primera fila del pallet
    cliente: str | None = None               # solo en formato TNLC
    certificado_grasp: bool = False


@dataclass
class ClienteVAT:
    nombre: str
    vat: str | None
    pallets: int | None
    cajas: int | None
    referencia: str | None


@dataclass
class PackingList:
    contenedor_codigo: str
    warehouse: str | None
    eta: dt.date | None
    fecha: dt.date | None
    total_pallets: int | None
    total_cajas: int | None
    pallets: list[PalletDetalle] = field(default_factory=list)
    clientes_vat: list[ClienteVAT] = field(default_factory=list)
    formato: str = "legacy"      # 'legacy' | 'tnlc'
    avisos: list[str] = field(default_factory=list)


# ───────────────────────── Helpers ───────────────────────────────────────────
def _solo_fecha(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def _f(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _i(v) -> int | None:
    if v in (None, ""):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _s(v) -> str | None:
    if v in (None, ""):
        return None
    return str(v).strip()


# ───────────────────────── TNLC nuevo formato ────────────────────────────────
def _parsear_tnlc(ws) -> PackingList:
    """Una sola hoja. Header en fila 8.

    R4-R6: cabecera (WAREHOUSE / ETA / EXPORTACIÓN # / CONTENEDOR / TOTAL PALLETS / TOTAL BOXES).
    R8-R9: encabezados de tabla principal y de la tabla cliente/VAT.
    R10..: filas de detalle pallets. Bloque cliente/VAT en cols J..N a partir
           de R11.
    """
    pl = PackingList(
        contenedor_codigo=_s(ws.cell(5, 2).value) or "",
        warehouse=_s(ws.cell(4, 2).value),
        eta=_solo_fecha(ws.cell(4, 5).value),
        fecha=None,
        total_pallets=_i(ws.cell(6, 2).value),
        total_cajas=_i(ws.cell(6, 5).value),
        formato="tnlc",
    )

    # Detalle de pallets
    fila = 10
    pallet_actual: int | None = None
    while True:
        v = ws.cell(fila, 1).value
        if v in (None, "") and ws.cell(fila, 2).value in (None, ""):
            # Verificar dos vacías seguidas como fin de tabla
            if ws.cell(fila + 1, 1).value in (None, ""):
                break
            fila += 1
            continue
        no_p = _i(v) or pallet_actual
        if no_p is None:
            fila += 1
            continue
        pallet_actual = no_p
        pl.pallets.append(
            PalletDetalle(
                no_pallet=no_p,
                calibre=str(ws.cell(fila, 2).value or "").strip(),
                presentacion_caja=None,
                predio=_s(ws.cell(fila, 3).value),
                ggn=_s(ws.cell(fila, 4).value),
                ica=_s(ws.cell(fila, 5).value),
                cliente=_s(ws.cell(fila, 6).value),
                cajas=_f(ws.cell(fila, 7).value),
                no_cargue=0,                       # TNLC no trae no_cargue por pallet aquí
                certificado_grasp=bool(ws.cell(fila, 8).value) and
                                   str(ws.cell(fila, 8).value).strip().lower() in ("true", "si", "yes", "1"),
            )
        )
        fila += 1
        if fila > 6000:  # guardarraíl
            break

    # Bloque clientes/VAT (cols J..N a partir de R12 hasta encontrar 'TOTAL')
    fila = 12
    while True:
        vat = _s(ws.cell(fila, 10).value)
        nombre = _s(ws.cell(fila, 11).value)
        if not nombre:
            break
        if nombre.upper() == "TOTAL":
            break
        pl.clientes_vat.append(
            ClienteVAT(
                nombre=nombre,
                vat=vat,
                pallets=_i(ws.cell(fila, 12).value),
                cajas=_i(ws.cell(fila, 13).value),
                referencia=_s(ws.cell(fila, 14).value),
            )
        )
        fila += 1
        if fila > 6000:
            break

    return pl


# ───────────────────────── Legacy (3 hojas) ─────────────────────────────────
def _parsear_legacy(wb) -> PackingList:
    resumen = wb["RESUMEN CONTENEDOR"]
    detalle = wb["DETALLE POR PALLET"]

    contenedor = _s(resumen.cell(4, 3).value) or ""  # 'MAQUILA # 326'
    m = re.search(r"#\s*(\d+)", contenedor)
    if m:
        contenedor = f"OP-{m.group(1)}"

    fecha_inicio = _solo_fecha(resumen.cell(6, 4).value)
    # Validar que la fecha no sea absurdamente vieja (>1 año atrás del año
    # actual). Algunos packing lists vienen con año 2021 / 2022 por tipeo de
    # la maquila (ej. PACKING LIST EXP #327 = "20/06/2021" en lugar de 2026).
    # Si pasa, dejamos None: el cruce caerá a today() como referencia.
    if fecha_inicio and fecha_inicio.year < dt.date.today().year - 1:
        fecha_inicio = None
    total_2kg = _i(resumen.cell(11, 8).value)
    total_pallets = _i(resumen.cell(12, 8).value) or total_2kg

    pl = PackingList(
        contenedor_codigo=contenedor,
        warehouse=None,
        eta=None,
        fecha=fecha_inicio,
        total_pallets=total_pallets,
        total_cajas=None,
        formato="legacy",
    )

    # Detalle: encabezados en R4, datos desde R5
    fila = 5
    pallet_actual: int | None = None
    while True:
        v = detalle.cell(fila, 1).value
        if v in (None, ""):
            # Cuando el pallet ocupa varias filas, sólo la primera trae '# PALLET'
            if pallet_actual is None or detalle.cell(fila, 2).value in (None, ""):
                # Si dos seguidas vacías → fin
                if detalle.cell(fila + 1, 1).value in (None, "") and \
                   detalle.cell(fila + 1, 2).value in (None, ""):
                    break
                fila += 1
                continue
            no_p = pallet_actual
        else:
            no_p = _i(v) or pallet_actual
            if no_p is not None:
                pallet_actual = no_p

        if pallet_actual is None:
            fila += 1
            continue

        total_cajas_pal = _i(detalle.cell(fila, 9).value)
        pl.pallets.append(
            PalletDetalle(
                no_pallet=pallet_actual,
                calibre=str(detalle.cell(fila, 2).value or "").strip(),
                presentacion_caja=_s(detalle.cell(fila, 3).value),
                predio=_s(detalle.cell(fila, 4).value),
                ica=_s(detalle.cell(fila, 5).value),
                ggn=_s(detalle.cell(fila, 6).value),
                no_cargue=_i(detalle.cell(fila, 7).value) or 0,
                cajas=_f(detalle.cell(fila, 8).value),
                total_cajas_pallet=total_cajas_pal,
            )
        )
        fila += 1
        if fila > 6000:
            break

    return pl


# ───────────────────────── API pública ───────────────────────────────────────
def parsear(ruta: str | Path) -> PackingList:
    ruta = Path(ruta)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    if "RESUMEN CONTENEDOR" in wb.sheetnames and "DETALLE POR PALLET" in wb.sheetnames:
        return _parsear_legacy(wb)
    # Fallback: hoja única → TNLC
    return _parsear_tnlc(wb.active)


if __name__ == "__main__":  # pragma: no cover
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Uso: python -m ingesta.parser_packing <archivo.xlsx>")
    pl = parsear(sys.argv[1])
    print(json.dumps(
        {
            "contenedor": pl.contenedor_codigo,
            "formato": pl.formato,
            "warehouse": pl.warehouse,
            "eta": str(pl.eta) if pl.eta else None,
            "total_pallets": pl.total_pallets,
            "total_cajas": pl.total_cajas,
            "pallets_filas": len(pl.pallets),
            "clientes_vat": len(pl.clientes_vat),
        },
        indent=2,
        default=str,
    ))
