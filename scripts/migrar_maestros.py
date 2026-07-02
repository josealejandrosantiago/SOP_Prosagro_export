"""Migración de maestros desde los Excel a Postgres:
  - precio_fruta       (Base de datos gulupa.xlsx / 'Precio de compra fruta')
  - clientes           (Costos consolidados.xlsx / 'CLIENTES ACTUALES')
  - proveedores        (Costos consolidados.xlsx / 'BD Proveedores')

Después de cargar precio_fruta, reconstruye kg_consolidado con el motor SOP
para que los costos dejen de ser 0.

Mapeos verificados con el workflow de análisis (01/07/2026):
  precio_fruta: col1 zona(EXTERNA), col2 lote(TEXTO), col3/4 fechas,
    col6 precio_nal, col7 precio_desh, col10 precio_expo (¡NO col5 que trae GGN!),
    col12 dias_pago, col13 consolidar_canastillas, col14 pagar_canastillas.
  clientes: col2 nombre, col1 vat, col5 pais, col8 correo.
  proveedores: col1 nit(TEXTO), col2 nombre; dedup por NIT.

Uso:
  python -m scripts.migrar_maestros            # dry-run (no escribe)
  python -m scripts.migrar_maestros --no-dry-run
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except AttributeError:
    pass

from conexion import get_conn  # noqa: E402

GULUPA = (
    r"C:\Users\LENOVO\OneDrive - Grupo San Jose\Automatización Prosagro Export - "
    r"Base de Datos Prosagro Export 1\01. Prosagro Export\01. Entrada De Datos\Base de datos gulupa.xlsx"
)
COSTOS = (
    r"C:\Users\LENOVO\OneDrive - Grupo San Jose\Automatización Prosagro Export - "
    r"Base de Datos Prosagro Export 1\01. Prosagro Export\01. Entrada De Datos\Costos consolidados.xlsx"
)


# ───────────────────────── helpers ──────────────────────────────────────────
def _s(v) -> str | None:
    if v is None or v == "":
        return None
    return str(v).strip() or None


def _num(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _int(v):
    if v in (None, ""):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _fecha(v):
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _bool_si(v) -> bool:
    return str(v).strip().lower() == "si" if v is not None else False


def _norm_lote(v) -> str:
    """Normaliza el lote igual que el parser de maquila: zfill(2) si es numérico.
    Así '6'→'06', 45→'45', '04'→'04', '101'→'101'. Crítico para que matchee
    con ingresos.lote y no deje el costo en 0."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v)).zfill(2)
    s = str(v).strip()
    return s.zfill(2) if s.isdigit() else s


def _mapa_zona_ext_int() -> dict[str, str]:
    with get_conn() as c, c.cursor() as cur:
        cur.execute("SELECT codigo_externo, codigo_interno FROM prosagro.zonas")
        return {str(ext): interno for ext, interno in cur.fetchall()}


# ───────────────────────── precio_fruta ─────────────────────────────────────
def cargar_precio_fruta(dry: bool) -> dict:
    wb = openpyxl.load_workbook(GULUPA, data_only=True, read_only=True)
    ws = wb["Precio de compra fruta"]
    zmap = _mapa_zona_ext_int()

    filas = []
    skip_zona = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        zona_ext = str(row[0]).strip().replace(".0", "")
        zona_int = zmap.get(zona_ext)
        if zona_int is None:
            skip_zona += 1
            continue
        filas.append({
            "zona": zona_int,
            "lote": _norm_lote(row[1]),
            "desde": _fecha(row[2]) or dt.date(2024, 1, 1),
            "hasta": _fecha(row[3]),
            "precio_nal": _num(row[5]),          # col6
            "precio_desh": _num(row[6]),         # col7
            "precio_expo": _num(row[9]),         # col10  ← Costo Exportación
            "dias_pago": _int(row[11]) or 0,     # col12
            "consolidar": _bool_si(row[12]),     # col13
            "pagar_can": _bool_si(row[13]),      # col14
            "moneda": "COP",
        })
    wb.close()

    print(f"  precio_fruta: {len(filas)} filas leídas (saltadas zona desconocida: {skip_zona})")
    print(f"    muestra: {filas[0] if filas else '(vacío)'}")
    if dry or not filas:
        return {"precio_fruta": len(filas)}

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("TRUNCATE prosagro.precio_fruta RESTART IDENTITY")
        cur.executemany(
            """
            INSERT INTO prosagro.precio_fruta
                (zona, lote, fecha_vigencia_desde, fecha_vigencia_hasta,
                 precio_expo, precio_nal, precio_desh, dias_pago,
                 consolidar_canastillas, pagar_canastillas, moneda)
            VALUES (%(zona)s, %(lote)s, %(desde)s, %(hasta)s,
                    %(precio_expo)s, %(precio_nal)s, %(precio_desh)s, %(dias_pago)s,
                    %(consolidar)s, %(pagar_can)s, %(moneda)s)
            """,
            filas,
        )
        conn.commit()
    print(f"  ✓ precio_fruta: {len(filas)} filas cargadas")
    return {"precio_fruta": len(filas)}


# ───────────────────────── clientes ─────────────────────────────────────────
def _norm_pais(v: str | None) -> str | None:
    if not v:
        return None
    s = v.strip().upper()
    if s in ("HOLANDA", "PAISES BAJOS", "PAÍSES BAJOS"):
        return "Países Bajos"
    return v.strip().title()


def cargar_clientes(dry: bool) -> dict:
    wb = openpyxl.load_workbook(COSTOS, data_only=True, read_only=True)
    ws = wb["CLIENTES ACTUALES"]
    filas = []
    vistos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = _s(row[1])  # col2 RAZON SOCIAL
        if not nombre or nombre.upper() in vistos:
            continue
        vistos.add(nombre.upper())
        vat = _s(row[0])
        if vat:
            vat = vat.replace(" ", "")
        correo = _s(row[7])
        if correo:
            # tomar el primer correo si hay varios separados por ; / espacio
            for sep in (";", ",", " "):
                if sep in correo:
                    correo = correo.split(sep)[0].strip()
                    break
        filas.append({
            "nombre": nombre,
            "vat": vat,
            "pais": _norm_pais(_s(row[4])),
            "correo": correo,
            "activo": True,
        })
    wb.close()

    print(f"  clientes: {len(filas)} únicos")
    print(f"    muestra: {filas[0] if filas else '(vacío)'}")
    if dry or not filas:
        return {"clientes": len(filas)}

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("TRUNCATE prosagro.clientes RESTART IDENTITY CASCADE")
        cur.executemany(
            """
            INSERT INTO prosagro.clientes (nombre, vat, pais, correo, activo)
            VALUES (%(nombre)s, %(vat)s, %(pais)s, %(correo)s, %(activo)s)
            """,
            filas,
        )
        conn.commit()
    print(f"  ✓ clientes: {len(filas)} cargados")
    return {"clientes": len(filas)}


# ───────────────────────── proveedores ──────────────────────────────────────
def _tipo_proveedor(nombre: str) -> str:
    up = nombre.upper()
    juridica = any(t in up for t in (" SAS", " S.A.S", " S.A", " SA ", " LTDA", " S.A.", "S.A.S.", "SAS"))
    return "JURIDICA" if juridica else "NATURAL"


def cargar_proveedores(dry: bool) -> dict:
    wb = openpyxl.load_workbook(COSTOS, data_only=True, read_only=True)
    ws = wb["BD Proveedores"]
    por_nit: dict[str, dict] = {}
    colisiones = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nit = _s(row[0])
        nombre = _s(row[1])
        if not nit or not nombre:
            continue
        nit = nit.replace(".0", "").split("-")[0].strip()
        if nit in por_nit:
            if por_nit[nit]["nombre"].upper() != nombre.upper():
                colisiones.append((nit, por_nit[nit]["nombre"], nombre))
            continue
        por_nit[nit] = {"nit": nit, "nombre": nombre, "tipo": _tipo_proveedor(nombre), "activo": True}
    wb.close()

    filas = list(por_nit.values())
    print(f"  proveedores: {len(filas)} NITs únicos")
    if colisiones:
        print(f"    ⚠ {len(colisiones)} NITs con distinto nombre (se conservó el primero):")
        for nit, n1, n2 in colisiones:
            print(f"       {nit}: '{n1}' vs '{n2}'")
    if dry or not filas:
        return {"proveedores": len(filas)}

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("TRUNCATE prosagro.proveedores RESTART IDENTITY CASCADE")
        cur.executemany(
            """
            INSERT INTO prosagro.proveedores (nit, nombre, tipo, activo)
            VALUES (%(nit)s, %(nombre)s, %(tipo)s, %(activo)s)
            """,
            filas,
        )
        conn.commit()
    print(f"  ✓ proveedores: {len(filas)} cargados")
    return {"proveedores": len(filas)}


# ───────────────────────── main ─────────────────────────────────────────────
def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--dry-run", action="store_true", default=True)
    p.add_argument("--no-dry-run", dest="dry_run", action="store_false")
    args = p.parse_args()
    dry = args.dry_run

    print("=" * 60)
    print("MIGRACIÓN DE MAESTROS" + ("  (DRY-RUN)" if dry else "  (ESCRIBIENDO)"))
    print("=" * 60)
    print("\n[1] precio_fruta")
    cargar_precio_fruta(dry)
    print("\n[2] clientes")
    cargar_clientes(dry)
    print("\n[3] proveedores")
    cargar_proveedores(dry)

    if not dry:
        print("\n[4] Reconstruir kg_consolidado (todas las semanas)")
        from app.servicios import sop_engine
        with get_conn() as c, c.cursor() as cur:
            cur.execute("SELECT DISTINCT anio, semana FROM prosagro.ingresos ORDER BY anio, semana")
            semanas = cur.fetchall()
        total = 0
        for anio, sem in semanas:
            r = sop_engine.reconstruir_kg_consolidado(anio=anio, semana=sem)
            total += r["procesadas"]
        print(f"  ✓ kg_consolidado reconstruido: {total} trazas en {len(semanas)} semanas")

        # Verificar costos en 0
        with get_conn() as c, c.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM prosagro.kg_consolidado
                WHERE costo_total_expo = 0 AND kg_expo_real > 0
            """)
            con_costo_0 = cur.fetchone()[0]
        print(f"  Trazas con kg_expo>0 pero costo_expo=0: {con_costo_0} (idealmente pocas)")
    else:
        print("\n(dry-run) — no se escribió a BD. Correr con --no-dry-run para aplicar.")


if __name__ == "__main__":
    main()
