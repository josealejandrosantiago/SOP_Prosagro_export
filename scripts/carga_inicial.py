"""CLI de carga inicial desde el Excel histórico de Prosagro.

Carga:
  - Maestros (productores, precio_fruta, precio_certificacion) desde
    `Base de datos gulupa.xlsx`.
  - Histórico de ingresos/export/nacional desde la misma hoja.
  - Histórico de contenedores y pallets desde los packing lists.

Uso:
    python -m scripts.carga_inicial maestros \\
        --gulupa "C:\\...\\Base de datos gulupa.xlsx"

    python -m scripts.carga_inicial historico \\
        --gulupa "C:\\...\\Base de datos gulupa.xlsx"

    python -m scripts.carga_inicial packing \\
        --carpeta "C:\\...\\Documentos\\Ingreso Gulupa\\operaciones\\2026"

Por seguridad: imprime un resumen y pide confirmación antes de tocar la BD.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

# Forzar UTF-8 en la consola Windows para que prints con acentos / símbolos
# como '→' o '✓' no truenen con cp1252.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except AttributeError:  # pragma: no cover
    pass

# Permitir `python scripts/carga_inicial.py` desde el root
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from conexion import get_conn  # noqa: E402


def _mapeo_zonas() -> dict[str, str]:
    """Lee tabla zonas y devuelve {codigo_externo: codigo_interno}.

    Se necesita porque el Excel viene en codigo_externo (122/123/124/125/126/127)
    pero la tabla productores referencia codigo_interno (01/02/03/04/05/06).
    """
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT codigo_externo, codigo_interno FROM prosagro.zonas")
        return {str(ext): inter for ext, inter in cur.fetchall()}


def _norm_txt(v) -> str | None:
    """None/'' → None; resto → string strippeado."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    return s or None


def _norm_telefono(v) -> str | None:
    """Excel a veces guarda teléfono como número científico — lo formateamos."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return f"{int(v):d}"
    return str(v).strip() or None


def cmd_maestros(args: argparse.Namespace) -> None:
    """Carga productores desde el Excel.

    El mapeo de columnas viene del header real de 'Base de datos gulupa.xlsx':
      col  1 (idx 0) — Fin vigencia
      col  2 (idx 1) — Inicio vigencia
      col  3 (idx 2) — Tipo movimiento (no se usa)
      col  4 (idx 3) — Zona (codigo_externo)
      col  5 (idx 4) — Lote
      col  6 (idx 5) — Estado (no se usa, ej. 'Cosechando' / 'Receso')
      col  9 (idx 8) — Nombre finca
      col 10 (idx 9) — Teléfono envío reportes producción
      col 12 (idx 11) — Ubicación_esp
      col 14 (idx 13) — Número de plantas
      col 17 (idx 16) — Nombre contacto (= propietario en VBA)
      col 18 (idx 17) — Identificación contacto (= documento en VBA)
      col 21 (idx 20) — ICA
      col 22 (idx 21) — GLOBAL (= GGN)
      col 27 (idx 26) — Retención
      col 28 (idx 27) — Facturación electrónica
    """
    wb = openpyxl.load_workbook(args.gulupa, data_only=True, read_only=True)

    nombre_hoja = next((s for s in wb.sheetnames if s.lower() == "productores"), None)
    if not nombre_hoja:
        print("  ✗ El Excel no tiene hoja 'Productores'")
        return
    ws = wb[nombre_hoja]

    mapeo = _mapeo_zonas()
    productores: list[dict] = []
    skip_zona = 0
    skip_propietario = 0
    skip_documento = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[3] is None:
            continue
        zona_ext = str(row[3]).strip()
        zona_int = mapeo.get(zona_ext)
        if zona_int is None:
            skip_zona += 1
            continue
        propietario = _norm_txt(row[16])
        documento = _norm_txt(row[17])
        if not propietario:
            skip_propietario += 1
            # OJO: aún así lo cargamos con placeholder para no perder el
            # registro (pueden ser lotes en setup) pero lo marcamos.
            propietario = "(sin propietario)"
        if not documento:
            skip_documento += 1
            documento = ""
        productores.append({
            "zona":             zona_int,
            "lote":             str(row[4] or "").strip(),
            "fecha_desde":      row[1],
            "fecha_hasta":      row[0],
            "nombre_finca":     _norm_txt(row[8]) or "(sin nombre)",
            "propietario":      propietario,
            "documento":        documento,
            "plantas":          row[13] if isinstance(row[13], (int, float)) else None,
            "ubicacion":        _norm_txt(row[11]),
            "telefono":         _norm_telefono(row[9]),
            "retencion":        _norm_txt(row[26]),
            "fact_electronica": _norm_txt(row[27]),
            "ica":              _norm_txt(row[20]),
            "ggn":              _norm_txt(row[21]),
        })

    print(f"  Productores leídos:           {len(productores)}")
    print(f"  Filas saltadas (zona desconocida): {skip_zona}")
    print(f"  Sin propietario (placeholder):     {skip_propietario}")
    print(f"  Sin documento (cadena vacía):      {skip_documento}")

    if args.dry_run or not productores:
        print("  (dry-run) — no se escribió a BD. Para escribir: --no-dry-run")
        # Pequeña muestra
        print("\n  Muestra (3 primeras filas):")
        for p in productores[:3]:
            print(f"    zona={p['zona']} lote={p['lote']} | finca={p['nombre_finca']!r}"
                  f" | propietario={p['propietario']!r} | doc={p['documento']!r}")
        return

    sql = """
        INSERT INTO prosagro.productores
            (zona, lote, fecha_vigencia_desde, fecha_vigencia_hasta,
             nombre_finca, propietario, documento, cantidad_plantas, ubicacion,
             telefono, requiere_retencion, facturacion_electronica,
             ica_propio, ggn_propio)
        VALUES (%(zona)s, %(lote)s, %(fecha_desde)s, %(fecha_hasta)s,
                %(nombre_finca)s, %(propietario)s, %(documento)s, %(plantas)s,
                %(ubicacion)s, %(telefono)s, %(retencion)s, %(fact_electronica)s,
                %(ica)s, %(ggn)s)
    """
    with get_conn() as conn, conn.cursor() as cur:
        cur.executemany(sql, productores)
        conn.commit()
    print(f"  ✓ {len(productores)} productores cargados")


def cmd_historico(args: argparse.Namespace) -> None:
    """Carga ingreso + export + nacional desde el libro consolidado."""
    print("  TODO Fase 1 — esqueleto disponible; carga real se implementa "
          "junto con el motor de Kg consolidado.")


def cmd_packing(args: argparse.Namespace) -> None:
    """Carga todos los packing lists de una carpeta."""
    from ingesta.parser_packing import parsear

    carpeta = Path(args.carpeta)
    archivos = sorted(carpeta.glob("PACKING LIST*.xlsx"))
    print(f"  Encontrados {len(archivos)} packing lists en {carpeta}")
    for f in archivos:
        try:
            pl = parsear(f)
            print(f"    {f.name:60s} → {pl.contenedor_codigo:10s} "
                  f"({pl.formato}, {len(pl.pallets)} filas, "
                  f"{len(pl.clientes_vat)} clientes)")
        except Exception as e:
            print(f"    {f.name:60s} → ERROR: {e}")

    if args.dry_run:
        print("  (dry-run) — no se escribió a BD.")
        return
    print("  TODO — escritura a BD se implementa en Fase 3 (Validación contenedor).")


def main() -> None:
    p = argparse.ArgumentParser(description="Carga inicial Prosagro Export")
    p.add_argument("--dry-run", action="store_true", default=True,
                   help="Por defecto NO escribe a BD. Quitar con --no-dry-run.")
    p.add_argument("--no-dry-run", dest="dry_run", action="store_false")
    sub = p.add_subparsers(dest="cmd", required=True)

    p1 = sub.add_parser("maestros")
    p1.add_argument("--gulupa", required=True,
                    help="Ruta al Excel 'Base de datos gulupa.xlsx'")
    p1.set_defaults(func=cmd_maestros)

    p2 = sub.add_parser("historico")
    p2.add_argument("--gulupa", required=True)
    p2.set_defaults(func=cmd_historico)

    p3 = sub.add_parser("packing")
    p3.add_argument("--carpeta", required=True)
    p3.set_defaults(func=cmd_packing)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
