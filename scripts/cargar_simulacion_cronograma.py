"""Carga a BD:
  - 'Simulación Viaje'      → prosagro.simulacion_viaje (volumen/incidencia/severidad)
  - 'Motivos rechazo fruta' → prosagro.causales_rechazo (incidencia histórica)
  - 'Cronograma'            → prosagro.cronograma_operaciones (BL/contenedor/invoice)
Todas del libro Base de datos gulupa.xlsx.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore
except AttributeError:
    pass

import openpyxl
from conexion import get_conn

GULUPA = (
    r"C:\Users\LENOVO\OneDrive - Grupo San Jose\Automatización Prosagro Export - "
    r"Base de Datos Prosagro Export 1\01. Prosagro Export\01. Entrada De Datos\Base de datos gulupa.xlsx"
)
ZONA_EXT_INT = {"122": "02", "123": "01", "124": "03", "125": "04", "126": "05", "127": "06"}


def _s(v):
    return None if v in (None, "") else str(v).strip()


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _fecha(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _zona_int(v):
    s = _s(v)
    if not s:
        return None
    s = s.replace(".0", "")
    return ZONA_EXT_INT.get(s, s)


def cargar():
    wb = openpyxl.load_workbook(GULUPA, data_only=True, read_only=True)
    resumen = {}

    with get_conn() as conn, conn.cursor() as cur:
        # ── Simulación Viaje ──────────────────────────────────────────────
        cur.execute("TRUNCATE prosagro.simulacion_viaje RESTART IDENTITY")
        ws = wb["Simulación Viaje"]
        sim = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            zona = _zona_int(row[6])   # col7 Zona
            if zona is None:
                continue
            fecha_elab = _fecha(row[0])
            anio = _int(row[4]) or (fecha_elab.year if fecha_elab else None)
            semana = _int(row[5])
            if anio is None or semana is None:
                continue
            sim.append((
                zona, _s(row[7]), anio, semana,
                _num(row[13]) or 0,          # col14 Volumen Export real
                _s(row[14]),                 # col15 Ubicación
                fecha_elab, _fecha(row[2]),  # fecha_elaboracion, fecha_inspeccion
                _s(row[3]),                  # col4 Tipo muestra
                _num(row[8]),                # col9 Cantidad muestra
                _s(row[9]),                  # col10 Evento
                _num(row[10]),               # col11 Cantidad evento
                _num(row[11]),               # col12 Porcentaje
                _num(row[20]),               # col21 Promedio Severidad
            ))
        cur.executemany(
            """INSERT INTO prosagro.simulacion_viaje
                (zona, lote, anio, semana, volumen, ubicacion, fecha_elaboracion,
                 fecha_inspeccion, tipo_muestra, cantidad_muestra, evento,
                 cantidad_evento, porcentaje, severidad_promedio)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            sim,
        )
        resumen["simulacion_viaje"] = len(sim)

        # ── Motivos rechazo fruta → causales_rechazo ──────────────────────
        cur.execute("TRUNCATE prosagro.causales_rechazo RESTART IDENTITY")
        ws = wb["Motivos rechazo fruta"]
        cau = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            traza = _s(row[2])          # col3 Codigo Trazabilidad
            causal = _s(row[3])         # col4 Causal
            if not traza or not causal:
                continue
            fecha = _fecha(row[0])
            pct = _num(row[4]) or 0     # col5 Porcentaje
            kg_nal = _num(row[6]) or 0  # col7 Kilos nacional
            zona = _zona_int(row[8])    # col9 Zona
            lote = _s(row[9])           # col10 Lote
            # severidad no viene explícita en esta hoja → clasificar por % (heurística)
            sev = "CRITICO" if pct >= 0.5 else ("MAYOR" if pct >= 0.2 else "MENOR")
            cau.append((fecha, traza, causal, pct, kg_nal, zona or "", lote or "",
                        sev, "Motivos rechazo fruta"))
        # insertar por lotes
        cur.executemany(
            """INSERT INTO prosagro.causales_rechazo
                (fecha, trazabilidad, causal, porcentaje, kg_nacional, zona, lote,
                 severidad, archivo_origen)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            cau,
        )
        resumen["causales"] = len(cau)

        # ── Cronograma ────────────────────────────────────────────────────
        cur.execute("TRUNCATE prosagro.cronograma_operaciones RESTART IDENTITY")
        ws = wb["Cronograma"]
        crono = []
        vistos = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            cod = _s(row[0])            # col1 Nro contenedor planta (OP-xxx)
            if not cod or cod in vistos:
                continue
            vistos.add(cod)
            crono.append((
                cod, _s(row[1]), _s(row[2]), _s(row[3]),
                _fecha(row[4]), _fecha(row[5]), _fecha(row[6]), _fecha(row[7]),
                _s(row[8]), _int(row[9]), _s(row[10]), _s(row[11]), _s(row[12]),
                _num(row[13]), _s(row[14]), _s(row[15]), _s(row[16]), _s(row[17]),
                _s(row[18]), _s(row[19]), _s(row[20]), _num(row[21]), _int(row[22]), _s(row[23]),
            ))
        cur.executemany(
            """INSERT INTO prosagro.cronograma_operaciones
                (contenedor_codigo, importador, invoice, puerto_origen,
                 fecha_salida_planta, fecha_embarque, fecha_llegada, fecha_buen_arribo,
                 puerto_destino, semana_llegada, contenedor_fisico, empresa_transporte,
                 vehiculo, tarifa_flete_terrestre, icoterms_real, observaciones,
                 booking, naviera, motonave, bl, maquila, dias_transito,
                 semana_salida, icoterm_facturacion)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
               ON CONFLICT (contenedor_codigo) DO NOTHING""",
            crono,
        )
        resumen["cronograma"] = len(crono)
        conn.commit()
    wb.close()
    return resumen


if __name__ == "__main__":
    r = cargar()
    for k, v in r.items():
        print(f"  {k}: {v}")
