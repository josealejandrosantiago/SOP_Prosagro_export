"""Cargador del histórico 2026 desde `Base de datos gulupa.xlsx` a BD.

Lee las 3 hojas operativas del Excel central (Ingreso Gulupa / Fruta
Exportación / Fruta Nacional) y persiste a las tablas prosagro.* sólo las
trazabilidades formato `2026 ...` (el formato canónico que usa la maquila a
partir del 2026). Las trazas con formato antiguo (`848 151 616 02 22`,
`505 144 616 02 32`, etc.) se ignoran porque corresponden a años anteriores y
no aplican al modelo nuevo.

Idempotente: UPSERT por trazabilidad en ingresos; reemplaza export/nacional
asociados a las trazas tocadas.

Uso:
    from app.servicios.historico_loader import cargar_excel_2026
    cargar_excel_2026(
        ruta_excel=...,
        anio_filtro=2026,
        semana_desde=1,
        semana_hasta=25,
    )
"""
from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterator

import openpyxl

from conexion import get_conn

TRAZA_2026_RE = re.compile(r"^2026\s+\d{3}\s+\d+\s+\d{2}\s+\d{1,3}$")
PARTS_RE = re.compile(r"\s+")


def _solo_fecha(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def _f(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _s(v) -> str | None:
    if v in (None, ""):
        return None
    return str(v).strip() or None


def _i(v) -> int | None:
    if v in (None, ""):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _parse_traza(traza: str) -> dict | None:
    """Parsea la trazabilidad. Soporta los DOS formatos históricos:
      - nuevo (2026): '2026 006 697 02 06' → no_cargue=697(parts[2]), zona=02, lote=06
      - viejo (2025): '505 144 616 02 32'  → no_cargue=505(parts[0]), zona=02, lote=32
    En AMBOS la zona está en parts[3] y el lote en parts[4]. El AÑO NO se saca
    de la traza (el formato viejo no lo trae) — se toma de la columna FECHA en
    el caller.
    """
    if not traza:
        return None
    parts = PARTS_RE.split(traza.strip())
    if len(parts) < 5:
        return None
    try:
        zona = parts[3]
        lote = parts[4]
        # formato nuevo: primer token es año de 4 dígitos → cargue en parts[2]
        if len(parts[0]) == 4 and parts[0].isdigit():
            no_cargue = int(parts[2])
        else:
            no_cargue = int(parts[0])
        return {"no_cargue": no_cargue, "zona": zona, "lote": lote}
    except (ValueError, IndexError):
        return None


def cargar_excel_2026(
    ruta_excel: str | Path,
    anios: set[int] | None = None,
    semana_desde: int = 1,
    semana_hasta: int = 53,
    excluir_ya_cargadas: bool = True,
    maquiladora_default: str = "Frutand",
) -> dict:
    """Carga histórico desde el Excel central a BD.

    Soporta AMBOS formatos de trazabilidad (nuevo '2026 ...' y viejo
    '505 144 616 02 32'). El AÑO se determina por la columna FECHA, no por la
    trazabilidad (el formato viejo no trae el año).

    - `anios`: conjunto de años a cargar (default {2025, 2026}).
    - `semana_desde/semana_hasta`: rango inclusivo.
    - `excluir_ya_cargadas`: si True, no recarga trazabilidades ya presentes.
    """
    if anios is None:
        anios = {2025, 2026}
    ruta_excel = Path(ruta_excel)
    print(f"  Abriendo {ruta_excel.name}…  (años: {sorted(anios)})")
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)

    # 1) Leer Ingreso Gulupa, Fruta Exportación, Fruta Nacional
    ingresos: list[dict] = []
    exports: list[dict] = []
    nacionales: list[dict] = []

    print("  Leyendo Ingreso Gulupa…")
    ws = wb["Ingreso Gulupa"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # col 16 (idx 15) Trazabilidad, col 17 (idx 16) Semana, col 5 (idx 4) FECHA INGRESO
        traza = _s(row[15])
        if not traza:
            continue
        p = _parse_traza(traza)
        if not p:
            continue
        fecha = _solo_fecha(row[4])
        if fecha is None or fecha.year not in anios:
            continue
        semana = _i(row[16]) or fecha.isocalendar()[1]
        if not (semana_desde <= semana <= semana_hasta):
            continue
        ingresos.append({
            "trazabilidad": traza,
            "semana": semana,
            "anio": fecha.year,
            "fecha_ingreso": fecha,
            "no_cargue": p["no_cargue"],
            "zona": p["zona"],
            "lote": p["lote"],
            "consec_int": f'{p["zona"]} {p["lote"]}',
            "canastillas": _i(row[12]) or 0,
            "peso_neto": _f(row[8]),
            "conductor": _s(row[13]),
            "placa": _s(row[14]),
            "finaliza": None,
        })
    print(f"    {len(ingresos)} ingresos en rango")

    print("  Leyendo Fruta Exportación…")
    ws = wb["Fruta Exportación"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # col 9 (idx 8) Trazabilidad, col 1 (idx 0) Semana, col 13 (idx 12) Calibre,
        # col 14 (idx 13) # Cajas empacadas, col 20 (idx 19) Kg empacados,
        # col 11 (idx 10) Fecha proceso, col 18 (idx 17) # Contenedor, col 26 (idx 25) Categoría
        traza = _s(row[8])
        if not traza:
            continue
        p = _parse_traza(traza)
        if not p:
            continue
        fecha_ingreso = _solo_fecha(row[9]) or _solo_fecha(row[10])
        fecha_proc = _solo_fecha(row[10]) or fecha_ingreso
        if fecha_ingreso is None or fecha_ingreso.year not in anios:
            continue
        semana = _i(row[0]) or fecha_ingreso.isocalendar()[1]
        if not (semana_desde <= semana <= semana_hasta):
            continue
        cal = _s(row[12]) or ""
        # El Excel a veces guarda calibre como número (38) o letra (M) — lo
        # normalizo a string upper.
        cal_norm = cal.upper() if isinstance(cal, str) else str(cal)
        exports.append({
            "trazabilidad": traza,
            "semana": semana,
            "anio": fecha_ingreso.year,
            "fecha_ingreso": fecha_ingreso,
            "fecha_procesamiento": fecha_proc,
            "no_cargue": p["no_cargue"],
            "dia_sem": None,
            "presentacion_caja": _s(row[20]),
            "calibre_desc": cal,
            "calibre_num": cal_norm,
            "id_calibre": None,
            "cant_cajas": _f(row[13]),
            "total_kg_netos": _f(row[19]),
            "productor_nombre": "PROSAGRO EXPORT SAS",
            "producto": "Gulupa",
            "ica": _s(row[6]),
            "ggn": _s(row[5]),
            "predio": _s(row[4]),
            "categoria": _s(row[25]) or "C1",
            "contenedor_codigo": _s(row[17]),
            "pallet_no": _i(row[15]),
            "armado_completo": _s(row[24]),
            "estado_cruce": "CRUZADO" if _s(row[17]) else "PENDIENTE",
        })
    print(f"    {len(exports)} fruta_export en rango")

    print("  Leyendo Fruta Nacional…")
    ws = wb["Fruta Nacional"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # col 1 (idx 0) FECHA INGRESO, col 2 (idx 1) TRAZABILIDAD, col 3 (idx 2) LOTE,
        # col 4 (idx 3) Fruta (kg descarte), col 6 (idx 5) Semana, col 7 (idx 6) C2
        traza = _s(row[1])
        if not traza:
            continue
        p = _parse_traza(traza)
        if not p:
            continue
        fecha = _solo_fecha(row[0])
        if fecha is None or fecha.year not in anios:
            continue
        semana = _i(row[5]) or fecha.isocalendar()[1]
        if not (semana_desde <= semana <= semana_hasta):
            continue
        nacionales.append({
            "trazabilidad": traza,
            "semana": semana,
            "anio": fecha.year,
            "fecha_ingreso": fecha,
            "fecha_procesamiento": fecha,
            "no_cargue": p["no_cargue"],
            "dia_sem": None,
            "lote_proceso": _s(row[2]),
            "merma": 0.0,
            "cant_kilos_descarte": _f(row[3]),
            "simulacion_kg": _f(row[6]),
        })
    print(f"    {len(nacionales)} fruta_nacional en rango")

    wb.close()

    if not ingresos:
        print("  Nada que cargar.")
        return {"ingresos": 0, "export": 0, "nacional": 0, "saltados": 0}

    # 2) Excluir trazas ya cargadas si excluir_ya_cargadas
    saltados = 0
    with get_conn() as conn, conn.cursor() as cur:
        if excluir_ya_cargadas:
            cur.execute(
                "SELECT trazabilidad FROM prosagro.ingresos WHERE trazabilidad = ANY(%s)",
                ([i["trazabilidad"] for i in ingresos],),
            )
            ya = {r[0] for r in cur.fetchall()}
            ingresos_n = [i for i in ingresos if i["trazabilidad"] not in ya]
            saltados = len(ingresos) - len(ingresos_n)
            if saltados:
                print(f"  Saltando {saltados} ingresos ya cargados a BD")
            ingresos = ingresos_n
        if not ingresos:
            return {"ingresos": 0, "export": 0, "nacional": 0, "saltados": saltados}

        trazas_a_cargar = {i["trazabilidad"] for i in ingresos}
        exports = [e for e in exports if e["trazabilidad"] in trazas_a_cargar]
        nacionales = [n for n in nacionales if n["trazabilidad"] in trazas_a_cargar]

        # 3) Insert ingresos
        print(f"  Persistiendo {len(ingresos)} ingresos…")
        cur.executemany(
            """
            INSERT INTO prosagro.ingresos
                (trazabilidad, semana, anio, fecha_ingreso, no_cargue, zona, lote,
                 consec_int, canastillas, peso_neto, conductor, placa, finaliza,
                 fruta_export_flag)
            VALUES (%(trazabilidad)s, %(semana)s, %(anio)s, %(fecha_ingreso)s,
                    %(no_cargue)s, %(zona)s, %(lote)s, %(consec_int)s,
                    %(canastillas)s, %(peso_neto)s, %(conductor)s, %(placa)s,
                    %(finaliza)s, TRUE)
            ON CONFLICT (trazabilidad) DO NOTHING
            """,
            ingresos,
        )

        # 4) Resolver trazabilidad → ingreso_id
        cur.execute(
            "SELECT trazabilidad, id FROM prosagro.ingresos WHERE trazabilidad = ANY(%s)",
            (list(trazas_a_cargar),),
        )
        traz_id = dict(cur.fetchall())

        # 5) Insert exports
        print(f"  Persistiendo {len(exports)} fruta_export…")
        cur.executemany(
            """
            INSERT INTO prosagro.fruta_export
                (ingreso_id, trazabilidad, semana, anio, fecha_ingreso,
                 fecha_procesamiento, no_cargue, presentacion_caja, calibre_desc,
                 calibre_num, id_calibre, cant_cajas, total_kg_netos,
                 productor_nombre, producto, ica, ggn, predio, categoria,
                 contenedor_codigo, pallet_no, armado_completo, estado_cruce)
            SELECT %(iid)s, %(trazabilidad)s, %(semana)s, %(anio)s, %(fecha_ingreso)s,
                   %(fecha_procesamiento)s, %(no_cargue)s, %(presentacion_caja)s,
                   %(calibre_desc)s, %(calibre_num)s, %(id_calibre)s, %(cant_cajas)s,
                   %(total_kg_netos)s, %(productor_nombre)s, %(producto)s,
                   %(ica)s, %(ggn)s, %(predio)s, %(categoria)s,
                   %(contenedor_codigo)s, %(pallet_no)s, %(armado_completo)s,
                   %(estado_cruce)s
            WHERE %(iid)s IS NOT NULL
            """,
            [{**e, "iid": traz_id.get(e["trazabilidad"])} for e in exports],
        )

        # 6) Insert nacional
        print(f"  Persistiendo {len(nacionales)} fruta_nacional…")
        cur.executemany(
            """
            INSERT INTO prosagro.fruta_nacional
                (ingreso_id, trazabilidad, semana, anio, fecha_ingreso,
                 fecha_procesamiento, no_cargue, lote_proceso, merma,
                 cant_kilos_descarte, simulacion_kg)
            SELECT %(iid)s, %(trazabilidad)s, %(semana)s, %(anio)s,
                   %(fecha_ingreso)s, %(fecha_procesamiento)s, %(no_cargue)s,
                   %(lote_proceso)s, %(merma)s, %(cant_kilos_descarte)s,
                   %(simulacion_kg)s
            WHERE %(iid)s IS NOT NULL
            """,
            [{**n, "iid": traz_id.get(n["trazabilidad"])} for n in nacionales],
        )

        conn.commit()

    return {
        "ingresos": len(ingresos),
        "export": len(exports),
        "nacional": len(nacionales),
        "saltados": saltados,
    }
